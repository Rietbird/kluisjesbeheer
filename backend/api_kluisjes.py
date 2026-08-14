import csv
import io
import zipfile
from flask import Blueprint, request, jsonify, g
from auth import login_required, beheerder_required, assert_vestiging_access, user_vestiging_ids
from klas import KLAS_SQL

kluisjes_bp = Blueprint('kluisjes', __name__, url_prefix='/api')

# Maximale gedecomprimeerde grootte van een XLSX-bestand (in bytes).
# XLSX is een zip, dus een 16MB-upload (MAX_CONTENT_LENGTH) kan tot
# GB's uitpakken (zip-bomb). 200MB is ruim voor een normaal kluis-
# bestand (3000 rijen ~3MB), klein genoeg om geheugen-DOS te voorkomen.
_MAX_XLSX_UNCOMPRESSED = 200 * 1024 * 1024


def _assert_kluisje_access(kid):
    """Kluisje-id -> vestiging -> access-check. None bij OK, Response bij fail."""
    row = g.db.execute('SELECT vestiging_id FROM kluisjes WHERE id = ?', (int(kid),)).fetchone()
    if not row:
        return jsonify({'error': 'Kluisje niet gevonden'}), 404
    return assert_vestiging_access(row['vestiging_id'])


def _assert_cluster_access(cid):
    """Cluster-id -> vestiging -> access-check. None bij OK, Response bij fail."""
    row = g.db.execute('SELECT vestiging_id FROM clusters WHERE id = ?', (int(cid),)).fetchone()
    if not row:
        return jsonify({'error': 'Cluster niet gevonden'}), 404
    return assert_vestiging_access(row['vestiging_id'])


def _safe_load_xlsx(file):
    """Open een XLSX met zip-bom-bescherming.
    Raised ValueError bij verdachte compressie-ratio's of total size.
    Caller moet file.seek(0) doen na deze check (file-pointer staat aan eind)."""
    import openpyxl
    # Stap 1: zip-content-check zonder uitpakken
    file.seek(0)
    try:
        with zipfile.ZipFile(file) as zf:
            total = sum(zi.file_size for zi in zf.infolist())
            if total > _MAX_XLSX_UNCOMPRESSED:
                raise ValueError(
                    f'XLSX zou na uitpakken {total // (1024*1024)}MB worden — '
                    f'maximum is {_MAX_XLSX_UNCOMPRESSED // (1024*1024)}MB '
                    f'(mogelijke zip-bomb).'
                )
    except zipfile.BadZipFile:
        raise ValueError('Bestand is geen geldige XLSX (zip-fout).')
    file.seek(0)
    return openpyxl.load_workbook(file, read_only=True)

@kluisjes_bp.route('/clusters/<int:cid>/kluisjes', methods=['GET'])
@login_required
def list_kluisjes(cid):
    err = _assert_cluster_access(cid)
    if err: return err
    rows = g.db.execute(
        'SELECT * FROM kluisjes WHERE cluster_id = ? AND verwijderd = 0 ORDER BY kluisnummer',
        (cid,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@kluisjes_bp.route('/kluisjes', methods=['GET'])
@login_required
def search_kluisjes():
    q = request.args.get('q', '').strip()
    vestiging_id = request.args.get('vestiging_id')
    status = request.args.get('status')
    klas = request.args.get('klas')

    query = f'''
        SELECT k.*, c.naam as cluster_naam, c.standaard_borg,
               t.id as toewijzing_id,
               t.leerling_naam, t.leerling_stamnr,
               {KLAS_SQL} AS leerling_klas,
               t.periode_van, t.periode_tot, t.borgbedrag, t.borg_betaald,
               t.reservesleutel_uitgegeven, t.reservesleutel_datum,
               l.vertrokken_op as leerling_vertrokken_op,
               l.nieuw_voor_schooljaar as leerling_nieuw_voor_schooljaar,
               CASE
                 WHEN k.status = 'vrij' AND EXISTS (
                   SELECT 1 FROM toewijzingen t2
                   WHERE t2.kluisje_id = k.id AND t2.actief = 0 AND t2.sleutel_ingeleverd = 0
                   AND t2.id = (SELECT MAX(t3.id) FROM toewijzingen t3 WHERE t3.kluisje_id = k.id)
                 ) THEN 1 ELSE 0
               END as _sleutel_niet_ingeleverd,
               CASE
                 WHEN k.status = 'vrij' AND EXISTS (
                   SELECT 1 FROM toewijzingen t2
                   WHERE t2.kluisje_id = k.id AND t2.actief = 0
                   AND t2.borg_betaald = 1 AND t2.borg_teruggestort = 0
                   AND t2.id = (SELECT MAX(t3.id) FROM toewijzingen t3 WHERE t3.kluisje_id = k.id)
                 ) THEN 1 ELSE 0
               END as _borg_niet_teruggestort
        FROM kluisjes k
        JOIN clusters c ON k.cluster_id = c.id
        LEFT JOIN toewijzingen t ON k.id = t.kluisje_id AND t.actief = 1
        LEFT JOIN leerlingen l ON t.leerling_stamnr = l.stamnr
        WHERE k.verwijderd = 0
    '''
    params = []

    if vestiging_id:
        err = assert_vestiging_access(vestiging_id)
        if err: return err
        query += ' AND k.vestiging_id = ?'
        params.append(int(vestiging_id))
    else:
        # Geen specifieke vestiging: filter op de user's allowed vestigingen
        # (beheerders zien alles, conciërges alleen hun eigen).
        allowed = user_vestiging_ids()
        if allowed is not None:
            if not allowed:
                return jsonify([])  # conciërge zonder vestigingen ziet niets
            placeholders = ','.join('?' * len(allowed))
            query += f' AND k.vestiging_id IN ({placeholders})'
            params.extend(allowed)
    # Sub-condities voor sleutel-statussen (hergebruikt door 'sleutel' = alle)
    _niet_ingeleverd = '''EXISTS (
            SELECT 1 FROM toewijzingen t2
            WHERE t2.kluisje_id = k.id AND t2.actief = 0 AND t2.sleutel_ingeleverd = 0
            AND t2.id = (SELECT MAX(t3.id) FROM toewijzingen t3 WHERE t3.kluisje_id = k.id AND t3.actief = 0)
        )'''
    _reservesleutel = '(t.actief = 1 AND t.reservesleutel_uitgegeven = 1)'
    if status == 'sleutel':
        # Alle sleutelkwesties: niet ingeleverd OF geen sleutel OF reservesleutel uitgegeven
        query += f' AND ({_niet_ingeleverd} OR k.geen_sleutel = 1 OR {_reservesleutel})'
    elif status == 'sleutel_niet_ingeleverd':
        query += f' AND {_niet_ingeleverd}'
    elif status == 'geen_sleutel':
        # Geen sleutel = vrij/buiten gebruik; een verhuurd kluisje hoort hier nooit
        query += " AND k.geen_sleutel = 1 AND k.status != 'uitgeleend'"
    elif status == 'reservesleutel':
        query += f' AND {_reservesleutel}'
    elif status == 'borg':
        # Lockers with outstanding borg: active with borg NOT paid, OR ended with borg paid but not refunded
        query += ''' AND (
            (t.actief = 1 AND t.borgbedrag > 0 AND t.borg_betaald = 0)
            OR EXISTS (
                SELECT 1 FROM toewijzingen t2
                WHERE t2.kluisje_id = k.id AND t2.actief = 0
                AND t2.borg_betaald = 1 AND t2.borg_teruggestort = 0
                AND t2.id = (SELECT MAX(t3.id) FROM toewijzingen t3 WHERE t3.kluisje_id = k.id)
            )
        )'''
    elif status == 'defect':
        # Defect is nu een aparte vlag, los van huurstatus
        query += ' AND k.is_defect = 1'
    elif status == 'vertrokken':
        # Bezet kluisje waarvan de huurder van school is (vertrokken_op gezet)
        query += ' AND t.actief = 1 AND l.vertrokken_op IS NOT NULL'
    elif status == 'vrij':
        # Vrij = écht uitleenbaar: geen kluisje-zonder-sleutel ertussen
        query += " AND k.status = 'vrij' AND k.geen_sleutel = 0"
    elif status:
        query += ' AND k.status = ?'
        params.append(status)
    if q:
        # Escape LIKE wildcards in user input
        q_escaped = q.replace('%', '\\%').replace('_', '\\_')
        query += """ AND (k.kluisnummer LIKE ? ESCAPE '\\' OR k.sleutelnummer LIKE ? ESCAPE '\\'
                     OR t.leerling_naam LIKE ? ESCAPE '\\' OR t.leerling_stamnr LIKE ? ESCAPE '\\')"""
        like = f'%{q_escaped}%'
        params.extend([like, like, like, like])
    if klas:
        query += f" AND {KLAS_SQL} = ?"
        params.append(klas)

    query += ' ORDER BY k.kluisnummer'
    rows = g.db.execute(query, params).fetchall()
    return jsonify([dict(r) for r in rows])

@kluisjes_bp.route('/vestigingen/<int:vid>/klassen', methods=['GET'])
@login_required
def klassen_in_vestiging(vid):
    """Unieke klassen met een actieve huurder in deze vestiging (voor de filter-dropdown)."""
    err = assert_vestiging_access(vid)
    if err: return err
    rows = g.db.execute(
        f'''SELECT DISTINCT {KLAS_SQL} AS klas
           FROM toewijzingen t
           JOIN kluisjes k ON t.kluisje_id = k.id
           LEFT JOIN leerlingen l ON t.leerling_stamnr = l.stamnr
           WHERE k.verwijderd = 0 AND t.actief = 1 AND k.vestiging_id = ?
             AND TRIM({KLAS_SQL}) <> ''
           ORDER BY klas''',
        (vid,)
    ).fetchall()
    return jsonify([r['klas'] for r in rows])


@kluisjes_bp.route('/kluisjes/<int:kid>', methods=['GET'])
@login_required
def get_kluisje(kid):
    err = _assert_kluisje_access(kid)
    if err: return err
    row = g.db.execute('SELECT * FROM kluisjes WHERE id = ? AND verwijderd = 0', (kid,)).fetchone()
    if not row:
        return jsonify({'error': 'Niet gevonden'}), 404
    return jsonify(dict(row))

@kluisjes_bp.route('/kluisjes/<int:kid>/sleutel-check', methods=['GET'])
@login_required
def sleutel_check(kid):
    """Advies-check: welke ANDERE kluisjes in dezelfde vestiging gebruiken dit
    sleutelnummer al? Blokkeert niets — sleutelnummers mogen dubbel zijn — maar
    laat de UI een waarschuwing tonen vóór opslaan."""
    err = _assert_kluisje_access(kid)
    if err: return err
    waarde = (request.args.get('waarde') or '').strip()
    row = g.db.execute('SELECT vestiging_id FROM kluisjes WHERE id = ? AND verwijderd = 0', (kid,)).fetchone()
    if not row:
        return jsonify({'error': 'Niet gevonden'}), 404
    if not waarde:
        return jsonify({'in_gebruik': False, 'kluisnummers': []})
    rows = g.db.execute(
        '''SELECT kluisnummer FROM kluisjes
           WHERE vestiging_id = ? AND TRIM(sleutelnummer) = ? AND verwijderd = 0 AND id != ?
           ORDER BY kluisnummer''',
        (row['vestiging_id'], waarde, kid)
    ).fetchall()
    return jsonify({'in_gebruik': len(rows) > 0, 'kluisnummers': [r['kluisnummer'] for r in rows]})


@kluisjes_bp.route('/clusters/<int:cid>/kluisjes', methods=['POST'])
@login_required
def create_kluisje(cid):
    err = _assert_cluster_access(cid)
    if err: return err
    data = request.get_json()
    kluisnummer = data.get('kluisnummer', '').strip()
    if not kluisnummer:
        return jsonify({'error': 'Kluisnummer is verplicht'}), 400

    cluster = g.db.execute('SELECT vestiging_id FROM clusters WHERE id = ?', (cid,)).fetchone()
    if not cluster:
        return jsonify({'error': 'Cluster niet gevonden'}), 404

    vestiging_id = cluster['vestiging_id']
    sleutelnummer = (data.get('sleutelnummer', '') or '').strip()
    locatie = data.get('locatie', '')

    try:
        cur = g.db.execute(
            'INSERT INTO kluisjes (cluster_id, vestiging_id, kluisnummer, sleutelnummer, locatie, status) VALUES (?, ?, ?, ?, ?, ?)',
            (cid, vestiging_id, kluisnummer, sleutelnummer, locatie, 'vrij')
        )
        g.db.commit()
    except Exception as e:
        if 'UNIQUE' in str(e):
            return jsonify({'error': f'Kluisnummer {kluisnummer} bestaat al in deze vestiging'}), 409
        raise

    row = g.db.execute('SELECT * FROM kluisjes WHERE id = ?', (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201

@kluisjes_bp.route('/kluisjes/<int:kid>', methods=['PUT'])
@login_required
def update_kluisje(kid):
    err = _assert_kluisje_access(kid)
    if err: return err
    data = request.get_json()
    row = g.db.execute('SELECT * FROM kluisjes WHERE id = ? AND verwijderd = 0', (kid,)).fetchone()
    if not row:
        return jsonify({'error': 'Niet gevonden'}), 404

    # Sleutelnummers zijn bewust NIET uniek (categorieën als 'Eraspas'/'Pasje'
    # en hergebruikte sleutels komen voor). De UI waarschuwt via /sleutel-check,
    # maar de backend blokkeert duplicaten niet.
    if 'sleutelnummer' in data:
        sleutelnummer = (data.get('sleutelnummer') or '').strip()
    else:
        sleutelnummer = row['sleutelnummer']
    locatie = data.get('locatie', row['locatie'])
    opmerkingen = data.get('opmerkingen', row['opmerkingen'])
    status = data.get('status', row['status'])
    if status not in ('vrij', 'uitgeleend'):
        # 'defect' als status mag niet meer (is_defect is nu apart veld)
        return jsonify({'error': 'Ongeldige status'}), 400

    # is_defect is een aparte vlag, los van huurstatus
    if 'is_defect' in data:
        is_defect = 1 if data.get('is_defect') else 0
        defect_sinds = row['defect_sinds']
        if is_defect and not row['is_defect']:
            defect_sinds = None  # zet onderstaand op datetime('now')
        elif not is_defect:
            defect_sinds = None
    else:
        is_defect = row['is_defect']
        defect_sinds = row['defect_sinds']

    # geen_sleutel is een aparte vlag (kluisje zonder bruikbare sleutel), los van huurstatus
    geen_sleutel = (1 if data.get('geen_sleutel') else 0) if 'geen_sleutel' in data else row['geen_sleutel']

    if is_defect and defect_sinds is None:
        g.db.execute(
            "UPDATE kluisjes SET sleutelnummer=?, locatie=?, opmerkingen=?, status=?, is_defect=?, defect_sinds=datetime('now'), geen_sleutel=?, updated_at=datetime('now') WHERE id=?",
            (sleutelnummer, locatie, opmerkingen, status, is_defect, geen_sleutel, kid)
        )
    else:
        g.db.execute(
            "UPDATE kluisjes SET sleutelnummer=?, locatie=?, opmerkingen=?, status=?, is_defect=?, defect_sinds=?, geen_sleutel=?, updated_at=datetime('now') WHERE id=?",
            (sleutelnummer, locatie, opmerkingen, status, is_defect, defect_sinds, geen_sleutel, kid)
        )
    g.db.commit()
    row = g.db.execute('SELECT * FROM kluisjes WHERE id = ?', (kid,)).fetchone()
    return jsonify(dict(row))

@kluisjes_bp.route('/kluisjes/<int:kid>', methods=['DELETE'])
@login_required
def delete_kluisje(kid):
    err = _assert_kluisje_access(kid)
    if err: return err
    row = g.db.execute('SELECT id FROM kluisjes WHERE id = ? AND verwijderd = 0', (kid,)).fetchone()
    if not row:
        return jsonify({'error': 'Niet gevonden'}), 404
    active = g.db.execute(
        'SELECT COUNT(*) as cnt FROM toewijzingen WHERE kluisje_id = ? AND actief = 1', (kid,)
    ).fetchone()['cnt']
    if active > 0:
        return jsonify({'error': 'Kan niet verwijderen: er is een actieve toewijzing'}), 409
    g.db.execute("UPDATE kluisjes SET verwijderd = 1, updated_at = datetime('now') WHERE id = ?", (kid,))
    g.db.commit()
    return jsonify({'ok': True})

@kluisjes_bp.route('/clusters/<int:cid>/kluisjes/bulk', methods=['POST'])
@login_required
def bulk_create_kluisjes(cid):
    """Bulk aanmaken van kluisjes. Body: { kluisjes: [{kluisnummer, sleutelnummer, locatie}, ...] }"""
    cluster = g.db.execute('SELECT vestiging_id FROM clusters WHERE id = ?', (cid,)).fetchone()
    if not cluster:
        return jsonify({'error': 'Cluster niet gevonden'}), 404
    err = assert_vestiging_access(cluster['vestiging_id'])
    if err: return err

    vestiging_id = cluster['vestiging_id']
    data = request.get_json() or {}
    items = data.get('kluisjes', [])
    if not items or len(items) > 500:
        return jsonify({'error': 'Geef 1–500 kluisjes op'}), 400

    created = []
    skipped = []
    for item in items:
        kluisnummer = str(item.get('kluisnummer', '')).strip()
        if not kluisnummer:
            continue
        sleutelnummer = str(item.get('sleutelnummer', '')).strip()
        locatie = str(item.get('locatie', '')).strip()
        try:
            cur = g.db.execute(
                'INSERT INTO kluisjes (cluster_id, vestiging_id, kluisnummer, sleutelnummer, locatie, status) VALUES (?, ?, ?, ?, ?, ?)',
                (cid, vestiging_id, kluisnummer, sleutelnummer, locatie, 'vrij')
            )
            created.append({'id': cur.lastrowid, 'kluisnummer': kluisnummer})
        except Exception as e:
            if 'UNIQUE' in str(e):
                skipped.append({'kluisnummer': kluisnummer, 'reden': 'Al bestaat'})
            else:
                g.db.rollback()
                raise

    g.db.commit()
    return jsonify({'created': len(created), 'skipped': skipped}), 201


@kluisjes_bp.route('/kluisjes/bulk-verwijderen', methods=['POST'])
@login_required
def bulk_delete_kluisjes():
    """Bulk verwijderen van kluisjes. Body: { kluisje_ids: [...] }"""
    data = request.get_json() or {}
    ids = data.get('kluisje_ids', [])
    if not ids or len(ids) > 500:
        return jsonify({'error': 'Geef 1–500 kluisje IDs op'}), 400

    deleted = 0
    skipped = []
    for kid in ids:
        row = g.db.execute('SELECT id, vestiging_id FROM kluisjes WHERE id = ? AND verwijderd = 0', (kid,)).fetchone()
        if not row:
            skipped.append({'kluisje_id': kid, 'reden': 'Niet gevonden'})
            continue
        err = assert_vestiging_access(row['vestiging_id'])
        if err:
            skipped.append({'kluisje_id': kid, 'reden': 'Geen toegang'})
            continue
        active = g.db.execute(
            'SELECT COUNT(*) as cnt FROM toewijzingen WHERE kluisje_id = ? AND actief = 1', (kid,)
        ).fetchone()['cnt']
        if active > 0:
            skipped.append({'kluisje_id': kid, 'reden': 'Actieve toewijzing'})
            continue
        g.db.execute("UPDATE kluisjes SET verwijderd = 1, updated_at = datetime('now') WHERE id = ?", (kid,))
        deleted += 1

    g.db.commit()
    return jsonify({'deleted': deleted, 'skipped': skipped})


def _parse_bedrag(text):
    """Parse '€ 10,00' or '10.00' to float."""
    if not text:
        return 0.0
    text = str(text).replace('\u20ac', '').replace('\ufffd', '').replace('€', '').strip()
    text = text.replace(',', '.').strip()
    try:
        return float(text)
    except (ValueError, TypeError):
        return 0.0


def _parse_periode_mx(text):
    """Parse MX format: 'van 1-8-2025 tot en met 31-7-2026' -> (iso_van, iso_tot).

    Accepts both d-m-yyyy and yyyy-m-d inside the 'van ... tot en met ...' string.
    """
    import re
    if not text:
        return None, None
    m = re.match(
        r'van\s+([\d:\- ]+?)\s+tot en met\s+([\d:\- ]+?|-)\s*$',
        str(text).strip()
    )
    if not m:
        return None, None
    return _parse_date_desktop(m.group(1)), _parse_date_desktop(m.group(2))


def _parse_date_desktop(text):
    """Parse a date to ISO (YYYY-MM-DD).

    Handles three inputs that Magister/Excel produce:
      - 'dd-mm-yyyy' or 'd-m-yyyy'           (string export)
      - '2025-08-01 00:00:00'                (str() of an Excel datetime)
      - a real datetime/date object
    """
    if text is None:
        return None
    # Real datetime/date object (openpyxl gives these for date cells)
    if hasattr(text, 'strftime'):
        return text.strftime('%Y-%m-%d')
    text = str(text).strip()
    if not text or text == '-':
        return None
    # Strip a trailing time component ('2025-08-01 00:00:00' or with 'T')
    text = text.split(' ')[0].split('T')[0].strip()
    parts = text.split('-')
    if len(parts) != 3:
        return text
    # ISO already (yyyy-mm-dd) -> normalise zero-padding, keep order
    if len(parts[0]) == 4:
        return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    # dd-mm-yyyy -> yyyy-mm-dd
    return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"


def _normaliseer_headers(row):
    """Lowercase header cells and collapse any whitespace to single spaces.

    Magister's Desktop export wraps headers over two lines, so the cell really
    contains 'Code\nKluisje'. Without collapsing that newline no lookup matches
    and every row is silently skipped.
    """
    import re
    return [re.sub(r'\s+', ' ', str(c or '')).strip().lower() for c in row]


def _detect_format(headers):
    """Detect whether the XLSX is Magister MX or Desktop format."""
    h_set = set(headers)
    if 'kluis' in h_set or 'uitleenperiode' in h_set or 'borgbedrag' in h_set:
        return 'mx'
    if ('omschrijving kluisje' in h_set or 'code kluisje' in h_set
            or 'verhuur vanaf' in h_set or 'stamnr' in h_set):
        return 'desktop'
    # Fallback: simple format (kluisnummer only)
    if 'kluisnummer' in h_set:
        return 'simple'
    return None


def _extract_prefix(kluisnummer):
    """Extract the vestiging-prefix from a locker number: the leading run of letters.

    The vestiging is encoded by the leading letter(s); everything from the first
    digit onward is the locker identifier (bank number, separator, sub-letter).
    So 'O053A' -> 'O', 'X001' -> 'X', 'Z01-A' -> 'Z', 'MO-0001' -> 'MO',
    'BL-001' -> 'BL'. A number that starts with a digit has no letter prefix
    and lands under 'Overig'.

    NB: this deliberately does NOT keep a trailing digit as part of the prefix
    (an earlier version did, for a hypothetical 'ISK1'). Doing so is ambiguous:
    'Z01-A' (vestiging Z) and 'ISK1-0003' (vestiging ISK1) are pattern-identical,
    and real data (Erasmus O/X/Z) needs the leading-letters rule.
    """
    import re
    m = re.match(r'^([A-Za-z]+)', kluisnummer)
    return m.group(1) if m else 'Overig'


def _normaliseer_kluisnummer(nummer, breedte):
    """Pad het eerste numerieke blok links met nullen tot `breedte`.

    Prefix en alles na het getal (suffix) blijven exact behouden.
    'MO-7' + 4 -> 'MO-0007'; 'MO-7B' + 4 -> 'MO-0007B';
    'BL-001' + 3 -> 'BL-001' (idempotent); zonder getal -> ongewijzigd.
    """
    import re
    if not nummer:
        return nummer
    m = re.match(r'^(.*?)(\d+)(.*)$', str(nummer))
    if not m:
        return nummer
    prefix, getal, rest = m.group(1), m.group(2), m.group(3)
    return f"{prefix}{getal.zfill(breedte)}{rest}"


def _analyseer_nummering(nummers):
    """Analyseer kluisnummers per prefix.

    Per prefix: is de cijferlengte variabel ('krom'), wat is de doelbreedte
    (cijfers van het hoogste getal), en zou normalisatie collisions geven.

    Let op: 'prefix' is hier alles vóór het eerste cijferblok, inclusief
    separator ('BL-001' -> 'BL-'). Dit wijkt bewust af van _extract_prefix()
    (dat 'BL' geeft); deze functie redeneert puur over normalisatie-breedte.
    """
    import re
    by_prefix = {}
    for nr in nummers:
        m = re.match(r'^(.*?)(\d+)(.*)$', str(nr or ''))
        if not m:
            continue
        prefix, getal, _ = m.group(1), m.group(2), m.group(3)
        d = by_prefix.setdefault(prefix, {
            'lengtes': set(), 'max': 0, 'nummers': []
        })
        d['lengtes'].add(len(getal))
        d['max'] = max(d['max'], int(getal))
        d['nummers'].append(nr)

    prefixes = []
    heeft_krom = False
    heeft_collision = False
    for prefix, d in sorted(by_prefix.items()):
        krom = len(d['lengtes']) > 1
        # Breedte minstens zo breed als het breedste bestaande nummer, anders
        # zou normalisatie cijfers afkappen (MO-001 -> MO-01) en info verliezen.
        breedte = max(len(str(d['max'])), max(d['lengtes']))
        # Collision: na normalisatie nog evenveel unieke nummers?
        genormaliseerd = {_normaliseer_kluisnummer(n, breedte) for n in d['nummers']}
        collision = len(genormaliseerd) != len(set(d['nummers']))
        if krom:
            heeft_krom = True
        if collision:
            heeft_collision = True
        prefixes.append({
            'prefix': prefix,
            'krom': krom,
            'breedte': breedte,
            'collision': collision,
            'aantal': len(d['nummers']),
        })
    return {
        'prefixes': prefixes,
        'heeft_krom': heeft_krom,
        'heeft_collision': heeft_collision,
    }


@kluisjes_bp.route('/kluisjes/import/preview', methods=['POST'])
@beheerder_required
def import_preview():
    """Scan an XLSX file and return a summary of prefixes, clusters, and locaties found."""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Bestand is verplicht'}), 400

    filename = file.filename or ''
    if not filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Alleen .xlsx bestanden worden geaccepteerd'}), 400

    try:
        wb = _safe_load_xlsx(file)
        ws = wb.active
        headers = None
        fmt = None
        prefixes = {}  # prefix -> count
        locaties = {}  # locatie -> count
        clusters = {}  # cluster -> count
        total = 0
        alle_nummers = []

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                headers = _normaliseer_headers(row)
                fmt = _detect_format(headers)
                if not fmt:
                    wb.close()
                    return jsonify({'error': f'Onbekend bestandsformaat'}), 400
                continue

            row_dict = dict(zip(headers, [str(c or '').strip() for c in row]))

            if fmt == 'mx':
                kluisnummer = row_dict.get('kluis', '')
                locatie = row_dict.get('locatie', '')
                cluster = row_dict.get('cluster', '')
            elif fmt == 'desktop':
                kluisnummer = row_dict.get('code kluisje', '') or row_dict.get('omschrijving kluisje', '')
                locatie = ''
                cluster = ''
            else:
                kluisnummer = row_dict.get('kluisnummer', '') or row_dict.get('kluis', '')
                locatie = ''
                cluster = ''

            if not kluisnummer:
                continue

            total += 1
            alle_nummers.append(kluisnummer)
            prefix = _extract_prefix(kluisnummer)
            prefixes[prefix] = prefixes.get(prefix, 0) + 1
            if locatie:
                locaties[locatie] = locaties.get(locatie, 0) + 1
            if cluster and cluster.lower() != 'zonder cluster':
                clusters[cluster] = clusters.get(cluster, 0) + 1

        wb.close()
    except Exception:
        return jsonify({'error': 'Kan bestand niet verwerken. Controleer het formaat.'}), 400

    return jsonify({
        'format': fmt,
        'total': total,
        'prefixes': [{'prefix': k, 'count': v} for k, v in sorted(prefixes.items())],
        'locaties': [{'locatie': k, 'count': v} for k, v in sorted(locaties.items())],
        'clusters': [{'cluster': k, 'count': v} for k, v in sorted(clusters.items())],
        'has_locaties': len(locaties) > 0,
        'normalisatie': _analyseer_nummering(alle_nummers),
    })


def _get_or_create_vestiging(naam):
    """Find existing vestiging by name, or create it."""
    row = g.db.execute('SELECT id FROM vestigingen WHERE naam = ?', (naam,)).fetchone()
    if row:
        return row['id']
    cur = g.db.execute('INSERT INTO vestigingen (naam) VALUES (?)', (naam,))
    return cur.lastrowid


def _get_or_create_cluster(vestiging_id, cluster_naam):
    """Find existing cluster by name + vestiging, or create it."""
    row = g.db.execute(
        'SELECT id FROM clusters WHERE vestiging_id = ? AND naam = ?',
        (vestiging_id, cluster_naam)
    ).fetchone()
    if row:
        return row['id']
    cur = g.db.execute(
        'INSERT INTO clusters (vestiging_id, naam) VALUES (?, ?)',
        (vestiging_id, cluster_naam)
    )
    return cur.lastrowid


@kluisjes_bp.route('/kluisjes/import', methods=['POST'])
@beheerder_required
def import_kluisjes():
    import json as json_mod
    cluster_id = request.form.get('cluster_id') or None
    vestiging_id = request.form.get('vestiging_id') or None
    auto_vestiging = request.form.get('auto_vestiging') == '1'
    normaliseer = request.form.get('normaliseer') == '1'
    # 'verhuur' matches rows against lockers that already exist and only writes
    # assignments. Needed because after the one-time migration every kluisnummer
    # is known, so the normal import skips every row and does nothing.
    modus = request.form.get('modus', '')
    beeindig_conflicten = request.form.get('beeindig_conflicten') == '1'
    # Dry run does the whole import and rolls back, so the preview numbers are
    # by construction the numbers of the real run.
    dry_run = request.form.get('dry_run') == '1'
    # prefix_mapping: JSON string {"BL": "Blauwlaken", "MO": "Molenstraat"}
    prefix_mapping = {}
    pm_raw = request.form.get('prefix_mapping', '')
    if pm_raw:
        try:
            prefix_mapping = json_mod.loads(pm_raw)
        except Exception:
            pass
    locatie_mapping = {}
    lm_raw = request.form.get('locatie_mapping', '')
    if lm_raw:
        try:
            locatie_mapping = json_mod.loads(lm_raw)
        except Exception:
            pass

    if not auto_vestiging and not vestiging_id and not prefix_mapping:
        if not cluster_id:
            return jsonify({'error': 'Kies een vestiging of gebruik automatisch uit bestand'}), 400
        cluster = g.db.execute('SELECT vestiging_id FROM clusters WHERE id = ?', (int(cluster_id),)).fetchone()
        if not cluster:
            return jsonify({'error': 'Cluster niet gevonden'}), 404
        vestiging_id = cluster['vestiging_id']
    elif vestiging_id:
        vestiging_id = int(vestiging_id)

    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Bestand is verplicht'}), 400

    filename = file.filename or ''
    if not filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Alleen .xlsx bestanden worden geaccepteerd'}), 400

    import re
    from datetime import date

    def _row_kluisnummer(fmt, row_dict):
        """Extract the kluisnummer string for a row, per format. Used in
        both the normalization pre-scan and the main import loop (DRY)."""
        if fmt == 'mx':
            return row_dict.get('kluis', '')
        elif fmt == 'desktop':
            # 'Code Kluisje' is the Verhuuroverzicht export, 'Omschrijving
            # Kluisje' the older kluisjes-overzicht export.
            return row_dict.get('code kluisje', '') or row_dict.get('omschrijving kluisje', '')
        else:  # simple
            return row_dict.get('kluisnummer', '') or row_dict.get('kluis', '')

    # Pre-scan: een read_only workbook-iterator is eenmalig, dus om de
    # doelbreedte per prefix te kennen lezen we alle kluisnummers in een
    # aparte pass. Daarna file.seek(0) + workbook opnieuw openen.
    breedte_per_prefix = {}
    if normaliseer:
        try:
            wb_scan = _safe_load_xlsx(file)
            ws_scan = wb_scan.active
            scan_headers = None
            scan_fmt = None
            scan_nummers = []
            for i, row in enumerate(ws_scan.iter_rows(values_only=True), start=1):
                if i == 1:
                    scan_headers = _normaliseer_headers(row)
                    scan_fmt = _detect_format(scan_headers)
                    if not scan_fmt:
                        break
                    continue
                row_dict = dict(zip(scan_headers, [str(c or '').strip() for c in row]))
                nr = _row_kluisnummer(scan_fmt, row_dict)
                if nr:
                    scan_nummers.append(nr)
            wb_scan.close()
            for p in _analyseer_nummering(scan_nummers)['prefixes']:
                breedte_per_prefix[p['prefix']] = p['breedte']
        except Exception:
            return jsonify({'error': 'Kan bestand niet verwerken. Controleer het formaat.'}), 400
        file.seek(0)

    try:
        wb = _safe_load_xlsx(file)
        ws = wb.active
        headers = None
        fmt = None
        kluisjes_created = 0
        toewijzingen_created = 0
        skipped = 0
        # verhuur-modus tellers
        toegewezen = ongewijzigd = conflicten = beeindigd = onbekend = 0

        def _maak_toewijzing(kluisje_id, stamnr, naam, klas, van, tot, borg):
            g.db.execute('''
                INSERT INTO toewijzingen
                (kluisje_id, leerling_stamnr, leerling_naam, leerling_klas,
                 periode_van, periode_tot, borgbedrag, borg_betaald, actief, aangemaakt_door)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            ''', (
                kluisje_id, stamnr, naam, klas, van, tot, borg,
                1 if borg > 0 else 0, 'XLSX Import',
            ))

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                headers = _normaliseer_headers(row)
                fmt = _detect_format(headers)
                if not fmt:
                    wb.close()
                    return jsonify({'error': f'Onbekend bestandsformaat. Verwachte kolommen niet gevonden. Gevonden: {", ".join(headers[:5])}...'}), 400
                continue

            row_dict = dict(zip(headers, [str(c or '').strip() for c in row]))

            kluisnummer = _row_kluisnummer(fmt, row_dict)

            if fmt == 'mx':
                sleutelnummer = row_dict.get('sleutel', '')
                locatie = row_dict.get('locatie', '')
                status_text = row_dict.get('status', '')
                naam = row_dict.get('naam', '')
                stamnr = row_dict.get('stamnummer', '')
                klas = row_dict.get('klas', '')
                borg = _parse_bedrag(row_dict.get('borgbedrag', ''))
                datum_van, datum_tot = _parse_periode_mx(row_dict.get('uitleenperiode', ''))
                opmerkingen = row_dict.get('opmerkingen', '')
                is_uitgeleend = status_text.lower() == 'uitgeleend'
                is_defect = status_text.lower() == 'defect'

            elif fmt == 'desktop':
                sleutelnummer = row_dict.get('sleutelnummer', '') or row_dict.get('slotnummer', '')
                locatie = ''
                stamnr = row_dict.get('stamnr', '')
                # Verhuuroverzicht has one 'Leerling' column; the older export
                # splits the name over three columns.
                naam = row_dict.get('leerling', '')
                if not naam:
                    achternaam = row_dict.get('achternaam', '')
                    tussenv = row_dict.get('tussenv', '')
                    roepnaam = row_dict.get('roepnaam', '')
                    naam = f"{roepnaam} {tussenv} {achternaam}".replace('  ', ' ').strip() if achternaam else ''
                klas = row_dict.get('klas', '')
                borg = 0.0
                datum_van = _parse_date_desktop(row_dict.get('verhuur vanaf', ''))
                datum_tot = _parse_date_desktop(row_dict.get('verhuur tot/met', ''))
                opmerkingen = ''
                is_uitgeleend = bool(naam and stamnr)
                is_defect = False

            else:  # simple
                sleutelnummer = row_dict.get('sleutelnummer', '') or row_dict.get('sleutel', '')
                locatie = row_dict.get('locatie', '')
                naam = ''
                stamnr = ''
                klas = ''
                borg = 0.0
                datum_van = None
                datum_tot = None
                opmerkingen = ''
                is_uitgeleend = False
                is_defect = False

            if not kluisnummer:
                continue

            # Normaliseer kluisnummer (zelfde prefix-definitie als
            # _analyseer_nummering: alles vóór het eerste cijferblok).
            if normaliseer:
                m_norm = re.match(r'^(.*?)(\d+)(.*)$', str(kluisnummer))
                if m_norm:
                    nprefix = m_norm.group(1)
                    nbreedte = breedte_per_prefix.get(nprefix)
                    if nbreedte:
                        kluisnummer = _normaliseer_kluisnummer(kluisnummer, nbreedte)

            # Determine vestiging for this row
            if prefix_mapping:
                prefix = _extract_prefix(kluisnummer)
                vest_naam = prefix_mapping.get(prefix, prefix_mapping.get('_default', ''))
                if vest_naam:
                    row_vestiging_id = _get_or_create_vestiging(vest_naam)
                elif vestiging_id:
                    row_vestiging_id = int(vestiging_id)
                else:
                    row_vestiging_id = _get_or_create_vestiging(prefix)
            elif auto_vestiging:
                row_locatie = row_dict.get('locatie', '').strip() if fmt == 'mx' else ''
                if row_locatie:
                    mapped = locatie_mapping.get(row_locatie, '').strip() if locatie_mapping else ''
                    row_vestiging_id = _get_or_create_vestiging(mapped or row_locatie)
                elif vestiging_id:
                    row_vestiging_id = int(vestiging_id)
                else:
                    row_vestiging_id = _get_or_create_vestiging('Onbekend')
            elif vestiging_id:
                row_vestiging_id = int(vestiging_id)
            else:
                row_vestiging_id = _get_or_create_vestiging('Onbekend')

            # Determine cluster for this row
            row_cluster_naam = row_dict.get('cluster', '').strip() if fmt == 'mx' else ''
            if row_cluster_naam and row_cluster_naam.lower() != 'zonder cluster':
                row_cluster_id = _get_or_create_cluster(row_vestiging_id, row_cluster_naam)
            elif cluster_id:
                row_cluster_id = int(cluster_id)
            else:
                row_cluster_id = _get_or_create_cluster(row_vestiging_id, 'Standaard')

            # Determine status (defect is een aparte vlag, los van huurstatus)
            db_status = 'uitgeleend' if is_uitgeleend else 'vrij'
            db_is_defect = 1 if is_defect else 0

            # Check for duplicate kluisnummer in this vestiging
            existing = g.db.execute(
                'SELECT id FROM kluisjes WHERE kluisnummer = ? AND vestiging_id = ? AND verwijderd = 0',
                (kluisnummer, row_vestiging_id)
            ).fetchone()

            if modus == 'verhuur':
                if not existing:
                    onbekend += 1
                    continue
                # A row without a student is NOT proof the locker is free:
                # Magister's locker administration is no longer maintained, so
                # a blank row must never end a running huur.
                if not (is_uitgeleend and naam):
                    continue
                if not datum_van:
                    datum_van = date.today().isoformat()
                if not datum_tot:
                    datum_tot = datum_van

                huidig = g.db.execute(
                    'SELECT id, leerling_stamnr FROM toewijzingen WHERE kluisje_id = ? AND actief = 1',
                    (existing['id'],)
                ).fetchone()
                if huidig and str(huidig['leerling_stamnr']) == str(stamnr):
                    ongewijzigd += 1
                    continue
                if huidig:
                    if not beeindig_conflicten:
                        conflicten += 1
                        continue
                    # The locker could only be handed to someone else because
                    # the previous key was returned, so record it as returned.
                    g.db.execute('''
                        UPDATE toewijzingen SET actief = 0, sleutel_ingeleverd = 1,
                        einddatum = ?, updated_at = datetime('now') WHERE id = ?
                    ''', (datum_van, huidig['id']))
                    beeindigd += 1

                _maak_toewijzing(existing['id'], stamnr, naam, klas, datum_van, datum_tot, borg)
                g.db.execute("UPDATE kluisjes SET status = 'uitgeleend', updated_at = datetime('now') WHERE id = ?",
                             (existing['id'],))
                toegewezen += 1
                continue

            if existing:
                skipped += 1
                continue

            # Insert kluisje
            cur = g.db.execute(
                "INSERT INTO kluisjes (cluster_id, vestiging_id, kluisnummer, sleutelnummer, locatie, status, is_defect, defect_sinds, opmerkingen) VALUES (?, ?, ?, ?, ?, ?, ?, CASE WHEN ?=1 THEN datetime('now') ELSE NULL END, ?)",
                (row_cluster_id, row_vestiging_id, kluisnummer, sleutelnummer, locatie, db_status, db_is_defect, db_is_defect, opmerkingen)
            )
            kluisjes_created += 1

            # Create toewijzing if uitgeleend
            if is_uitgeleend and naam:
                kluisje_id = cur.lastrowid
                if not datum_van:
                    datum_van = date.today().isoformat()
                if not datum_tot:
                    datum_tot = datum_van
                _maak_toewijzing(kluisje_id, stamnr, naam, klas, datum_van, datum_tot, borg)
                toewijzingen_created += 1

        wb.close()
        if dry_run:
            g.db.rollback()
        else:
            g.db.commit()
    except Exception as e:
        g.db.rollback()
        if 'UNIQUE' in str(e):
            return jsonify({'error': 'Duplicaat kluisnummer gevonden — import afgebroken'}), 400
        raise

    if modus == 'verhuur':
        return jsonify({
            'modus': 'verhuur',
            'toegewezen': toegewezen,
            'ongewijzigd': ongewijzigd,
            'conflicten': conflicten,
            'beeindigd': beeindigd,
            'onbekend': onbekend,
            'format': fmt,
        }), 201

    return jsonify({
        'imported': kluisjes_created,
        'toewijzingen': toewijzingen_created,
        'skipped': skipped,
        'format': fmt,
    }), 201
