#!/usr/bin/env python3
"""Generate example XLSX import file based on real Magister export format."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kluisgegevens"

# Headers matching the real Magister export
headers = ['Cluster', 'Kluis', 'Naam', 'Stamnummer', 'Klas', 'Uitleenperiode', 'Status',
           'Borgbedrag', 'Huurbedrag', 'Locatie', 'Sleutel', 'Slot', 'Studie', 'Kluistype']
ws.append(headers)

header_fill = PatternFill(start_color='FF8200', end_color='FF8200', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='left')

# Example rows - mix of uitgeleend and vrij
rows = [
    ['Vleugel A', 'A001', 'Jan de Vries', '21001', '3A', 'van 1-1-2026 tot en met 31-7-2026', 'Uitgeleend',
     '€ 10,00', '€ 0,00', 'Hoofdlocatie', 'S2040', '', '', 'Leerlingkluisje'],
    ['Vleugel A', 'A002', 'Emma Bakker', '22002', '2B', 'van 1-1-2026 tot en met 31-7-2026', 'Uitgeleend',
     '€ 10,00', '€ 0,00', 'Hoofdlocatie', 'S2656', '', '', 'Leerlingkluisje'],
    ['Vleugel A', 'A003', '', '', '', 'van - tot en met -', 'Vrij',
     '€ 10,00', '€ 0,00', 'Hoofdlocatie', 'S2723', '', '', 'Leerlingkluisje'],
    ['Zonder cluster', 'B100', 'Sophie Jansen', '20100', '4A', 'van 1-8-2025 tot en met 31-7-2026', 'Uitgeleend',
     '€ 10,00', '€ 0,00', 'Dependance Noord', 'S0100', '', '', 'Leerlingkluisje'],
    ['Zonder cluster', 'B101', '', '', '', 'van - tot en met -', 'Vrij',
     '€ 10,00', '€ 0,00', 'Dependance Noord', 'S0101', '', '', 'Leerlingkluisje'],
]

for row in rows:
    ws.append(row)

# Auto-width columns
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

wb.save('voorbeeld-import.xlsx')
print("Created voorbeeld-import.xlsx")
